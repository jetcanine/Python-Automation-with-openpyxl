import openpyxl
inventory_file = openpyxl.load_workbook('inventory.xlsx')
content=inventory_file['Sheet1']

# Display Product having inventory less than 10
product_below_10={}

for product_row in range(2, content.max_row + 1):
    supplier_name= content.cell(product_row, 4).value
    inventory= content.cell(product_row, 2).value
    product_num=content.cell(product_row,1).value

    if inventory<10:
        product_below_10[product_num]=inventory
        
print(product_below_10)