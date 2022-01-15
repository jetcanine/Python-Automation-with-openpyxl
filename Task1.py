import openpyxl
inventory_file = openpyxl.load_workbook('inventory.xlsx')
content=inventory_file['Sheet1']

# List The Number of Products Provided by the Suppliers

product_of_suppliers = {}

for row in range(2, content.max_row +1):
    supplier_name = content.cell(row, 4).value
    # print(supplier_name)

    if supplier_name in product_of_suppliers:
        current_products = product_of_suppliers[supplier_name]
        product_of_suppliers[supplier_name] = current_products + 1

    else:
        print("New Supplier Added")
        product_of_suppliers[supplier_name] = 1

print(product_of_suppliers)