import openpyxl
inventory_file = openpyxl.load_workbook('inventory.xlsx')
content=inventory_file['Sheet1']

#Total Inventory/Content Value per Supplier

total_value_per_supplier={}

for product_row in range(2, content.max_row + 1):
    supplier_name= content.cell(product_row,4).value
    inventory= content.cell(product_row,2).value
    price= content.cell(product_row,3).value

    if supplier_name in total_value_per_supplier:
        current_total= total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name]=current_total+inventory*price
    else:
        total_value_per_supplier[supplier_name]= inventory*price

print(total_value_per_supplier)

