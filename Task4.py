import openpyxl
inventory_file = openpyxl.load_workbook('inventory.xlsx')
content=inventory_file['Sheet1']

# Adding Content in New Row
for product_row in range(2, content.max_row + 1):
    inventory = content.cell(product_row, 2).value
    price = content.cell(product_row, 3).value
    inventory_price = content.cell(product_row, 5)

    inventory_price.value = inventory * price

inventory_file.save("Inventory_with_total_value.xlsx")